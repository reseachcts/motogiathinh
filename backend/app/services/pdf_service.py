"""PDF and Excel generation service using WeasyPrint and openpyxl."""
from __future__ import annotations

import html
from datetime import date
from io import BytesIO

_GENDER = {"male": "Nam", "female": "Nữ", "other": "Khác"}
_STATUS = {"pending": "Chờ duyệt", "active": "Đang học", "suspended": "Tạm dừng", "completed": "Hoàn thành", "dropped": "Nghỉ học"}
_LEARN = {"not_started": "Chưa học", "passed": "Đạt", "failed": "Không đạt"}
_PLAN_ST = {"pending": "Chờ đóng", "partial": "Còn thiếu", "paid": "Đã đóng", "overdue": "Quá hạn", "waived": "Miễn giảm", "refunded": "Hoàn tiền"}
_METHOD = {"cash": "Tiền mặt", "bank_transfer": "Chuyển khoản", "momo": "MoMo", "zalopay": "ZaloPay"}
_MON = ["Tháng 1","Tháng 2","Tháng 3","Tháng 4","Tháng 5","Tháng 6","Tháng 7","Tháng 8","Tháng 9","Tháng 10","Tháng 11","Tháng 12"]


def _v(v: object) -> str:
    if v is None:
        return "—"
    return f"{int(round(float(v))):,}₫".replace(",", ".")


def _d(d: object) -> str:
    if not d:
        return "—"
    if isinstance(d, date):
        return d.strftime("%d/%m/%Y")
    s = str(d)[:10].split("-")
    return f"{s[2]}/{s[1]}/{s[0]}" if len(s) == 3 else str(d)[:10]


def _e(s: object) -> str:
    return html.escape(str(s or "—"))


_CSS = """
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: 'Noto Sans', 'DejaVu Sans', Arial, sans-serif; font-size: 10.5pt; color: #1a1a1a; padding: 24px 28px; }
.hdr { display: flex; justify-content: space-between; align-items: flex-start; border-bottom: 2.5px solid #1677ff; padding-bottom: 12px; margin-bottom: 18px; }
.school { font-size: 14pt; font-weight: 700; color: #1677ff; }
.sub { font-size: 9pt; color: #666; margin-top: 4px; }
h2 { font-size: 10pt; font-weight: 700; color: #333; text-transform: uppercase; letter-spacing: .06em; border-bottom: 1px solid #e8e8e8; padding-bottom: 4px; margin: 16px 0 8px; }
table { width: 100%; border-collapse: collapse; font-size: 9.5pt; margin-bottom: 4px; }
thead th { background: #1677ff; color: #fff; padding: 5px 8px; text-align: left; font-weight: 600; }
td { padding: 5px 8px; border-bottom: 1px solid #f0f0f0; vertical-align: top; }
tr:nth-child(even) td { background: #f7f9ff; }
tfoot td { font-weight: 700; border-top: 1.5px solid #ddd; }
.kpis { display: flex; gap: 10px; margin-bottom: 14px; }
.kpi { flex: 1; border: 1px solid #e8e8e8; border-radius: 6px; padding: 10px 12px; background: #fafbff; }
.kpi-l { font-size: 8.5pt; color: #888; margin-bottom: 3px; }
.kpi-v { font-size: 13pt; font-weight: 700; color: #1677ff; }
.red { color: #f5222d !important; }
.green { color: #52c41a !important; }
.foot { margin-top: 20px; border-top: 1px solid #eee; padding-top: 10px; font-size: 8.5pt; color: #aaa; display: flex; justify-content: space-between; }
.badge { display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 8.5pt; }
.bg { background: #f6ffed; color: #389e0d; border: 1px solid #b7eb8f; }
.bb { background: #e6f4ff; color: #0958d9; border: 1px solid #91caff; }
.br { background: #fff2f0; color: #cf1322; border: 1px solid #ffa39e; }
.bo { background: #fffbe6; color: #d48806; border: 1px solid #ffe58f; }
"""


def _doc(title: str, body: str) -> str:
    return f'<!DOCTYPE html><html lang="vi"><head><meta charset="utf-8"><title>{html.escape(title)}</title><style>{_CSS}</style></head><body>{body}</body></html>'


def _svg_bar(
    items: list[tuple[str, float]],
    color: str = "#1677ff",
    fmt=None,
    width: int = 520,
    height: int = 130,
) -> str:
    """Render a bar chart as inline SVG for WeasyPrint."""
    if not items:
        return ""
    values = [v for _, v in items]
    max_v = max(values) if any(v > 0 for v in values) else 1
    if fmt is None:
        fmt = lambda x: f"{x / 1_000_000:.1f}tr" if x >= 500_000 else str(int(x))

    pad_l, pad_r, pad_t, pad_b = 44, 8, 12, 22
    chart_w = width - pad_l - pad_r
    chart_h = height - pad_t - pad_b
    n = len(items)
    slot_w = chart_w / n
    bar_w = max(4, slot_w * 0.65)

    parts: list[str] = []
    # Grid lines + Y labels at 0%, 50%, 100%
    for pct in (0.0, 0.5, 1.0):
        y = pad_t + chart_h * (1 - pct)
        label = fmt(max_v * pct)
        parts.append(f'<text x="{pad_l - 3}" y="{y + 3:.1f}" text-anchor="end" font-size="7" fill="#888">{html.escape(label)}</text>')
        parts.append(f'<line x1="{pad_l}" y1="{y:.1f}" x2="{width - pad_r}" y2="{y:.1f}" stroke="#ddd" stroke-dasharray="3,2"/>')

    # Bars + X labels
    for i, (label, val) in enumerate(items):
        bar_h = max(2.0, (val / max_v) * chart_h) if max_v > 0 else 2.0
        cx = pad_l + slot_w * i + slot_w / 2
        bx = cx - bar_w / 2
        by = pad_t + chart_h - bar_h
        parts.append(f'<rect x="{bx:.1f}" y="{by:.1f}" width="{bar_w:.1f}" height="{bar_h:.1f}" fill="{color}" rx="2"/>')
        parts.append(f'<text x="{cx:.1f}" y="{pad_t + chart_h + 14}" text-anchor="middle" font-size="7" fill="#555">{html.escape(str(label))}</text>')
        if bar_h > 16 and val > 0:
            parts.append(f'<text x="{cx:.1f}" y="{by - 2:.1f}" text-anchor="middle" font-size="6.5" fill="{color}">{html.escape(fmt(val))}</text>')

    inner = "".join(parts)
    return f'<svg width="100%" viewBox="0 0 {width} {height}" xmlns="http://www.w3.org/2000/svg">{inner}</svg>'


def _to_pdf(html_str: str) -> BytesIO:
    from weasyprint import HTML
    buf = BytesIO(HTML(string=html_str).write_pdf())
    buf.seek(0)
    return buf


# ── Student resume ────────────────────────────────────────────────────────────

def generate_student_resume(
    student: dict,
    enrollments: list[dict],
    plans: list[dict],
    contacts: list[dict],
) -> BytesIO:
    name = student.get("ten_hoc_vien", "")
    status = student.get("trang_thai", "")
    badge = {"active": "bg", "completed": "bb", "pending": "bo"}.get(status, "br")
    addr = ", ".join(x for x in [student.get(k, "") for k in ["dia_chi", "phuong_xa", "quan_huyen", "tinh_thanh"]] if x) or "—"

    body = f"""
<div class="hdr">
  <div>
    <div class="school">MÔ TÔ GIA THỊNH — HỒ SƠ HỌC VIÊN</div>
    <div class="sub">Mã: {_e(student.get("ma_hoc_vien"))} &nbsp;·&nbsp; Ngày xuất: {date.today().strftime("%d/%m/%Y")}</div>
  </div>
  <div style="text-align:right">
    <span class="badge {badge}">{html.escape(_STATUS.get(status, status))}</span>
    <div class="sub" style="margin-top:6px">Bằng lái: <b>{_e(student.get("loai_bang_lai"))}</b></div>
  </div>
</div>

<h2>Thông tin cá nhân</h2>
<table>
  <tr><td width="28%"><b>Họ và tên</b></td><td>{_e(name)}</td><td width="28%"><b>Ngày sinh</b></td><td>{_d(student.get("ngay_sinh"))} ({html.escape(_GENDER.get(student.get("gioi_tinh", ""), "—"))})</td></tr>
  <tr><td><b>CCCD</b></td><td>{_e(student.get("cccd_number"))}</td><td><b>Điện thoại</b></td><td>{_e(student.get("so_dien_thoai"))}</td></tr>
  <tr><td><b>Email</b></td><td>{_e(student.get("dia_chi_email"))}</td><td><b>Zalo</b></td><td>{_e(student.get("zalo_number"))}</td></tr>
  <tr><td><b>Địa chỉ</b></td><td colspan="3">{html.escape(addr)}</td></tr>
  <tr><td><b>Ngày đăng ký</b></td><td>{_d(student.get("ngay_dang_ky"))}</td><td><b>SK hết hạn</b></td><td>{_d(student.get("health_cert_expiry"))}</td></tr>
</table>"""

    if student.get("ho_ten_nguoi_than"):
        body += f"""
<h2>Người thân / khẩn cấp</h2>
<table>
  <tr><td width="28%"><b>Họ tên</b></td><td>{_e(student.get("ho_ten_nguoi_than"))}</td><td width="28%"><b>Quan hệ</b></td><td>{_e(student.get("quan_he"))}</td></tr>
  <tr><td><b>Điện thoại</b></td><td colspan="3">{_e(student.get("sdt_nguoi_than"))}</td></tr>
</table>"""

    if contacts:
        rows = "".join(f"<tr><td>{_e(c.get('contact_name'))}</td><td>{_e(c.get('relation'))}</td><td>{_e(c.get('phone'))}</td></tr>" for c in contacts)
        body += f"<h2>Liên hệ bổ sung</h2><table><thead><tr><th>Họ tên</th><th>Quan hệ</th><th>Điện thoại</th></tr></thead><tbody>{rows}</tbody></table>"

    if enrollments:
        rows = ""
        for e in enrollments:
            lop = e.get("lop_hoc") or {}
            ct = lop.get("course_type") or {}
            rows += f"""<tr>
  <td>{_e(lop.get("ten_lop"))}</td><td>{_e(lop.get("ma_lop"))}</td><td>{_e(ct.get("loai_bang_lai", ""))}</td>
  <td>{_d(e.get("enrollment_date"))}</td>
  <td>{html.escape(_LEARN.get(str(e.get("ly_thuyet_status", "")), ""))}</td>
  <td>{html.escape(_LEARN.get(str(e.get("thuc_hanh_status", "")), ""))}</td>
  <td>{e.get("overall_progress", 0)}%</td><td>{_d(e.get("completion_date"))}</td></tr>"""
        body += f"""<h2>Lịch sử học ({len(enrollments)} lớp)</h2>
<table><thead><tr><th>Lớp học</th><th>Mã lớp</th><th>Bằng</th><th>Ngày ĐK</th><th>Lý thuyết</th><th>Thực hành</th><th>Tiến độ</th><th>Hoàn thành</th></tr></thead><tbody>{rows}</tbody></table>"""

    if plans:
        net = sum(float(p.get("net_amount", 0)) for p in plans)
        paid = sum(float(p.get("paid_amount", 0)) for p in plans)
        rem = sum(float(p.get("remaining_amount", 0)) for p in plans)
        rows = "".join(
            f"<tr><td>{html.escape(_PLAN_ST.get(str(p.get('payment_status','')),str(p.get('payment_status',''))))}</td>"
            f"<td style='text-align:right'>{_v(p.get('net_amount'))}</td>"
            f"<td style='text-align:right'>{_v(p.get('paid_amount'))}</td>"
            f"<td style='text-align:right'>{_v(p.get('remaining_amount'))}</td>"
            f"<td>{_d(p.get('due_date'))}</td></tr>"
            for p in plans
        )
        body += f"""<h2>Thanh toán</h2>
<div class="kpis">
  <div class="kpi"><div class="kpi-l">Thực thu</div><div class="kpi-v">{_v(net)}</div></div>
  <div class="kpi"><div class="kpi-l">Đã đóng</div><div class="kpi-v green">{_v(paid)}</div></div>
  <div class="kpi"><div class="kpi-l">Còn lại</div><div class="kpi-v {'red' if rem > 0 else 'green'}">{_v(rem)}</div></div>
</div>
<table><thead><tr><th>Trạng thái</th><th>Thực thu</th><th>Đã đóng</th><th>Còn lại</th><th>Hạn</th></tr></thead>
<tbody>{rows}</tbody>
<tfoot><tr><td>Tổng</td><td style='text-align:right'>{_v(net)}</td><td style='text-align:right'>{_v(paid)}</td><td style='text-align:right'>{_v(rem)}</td><td>—</td></tr></tfoot></table>"""

    body += f'<div class="foot"><span>Hệ thống Mô Tô Gia Thịnh</span><span>Ngày in: {date.today().strftime("%d/%m/%Y")}</span></div>'
    return _to_pdf(_doc(f"Hồ sơ {name}", body))


# ── Analytics / Statistics report ────────────────────────────────────────────

def generate_analytics_report(
    analytics: dict,
    revenue: list[dict],
    year: int,
    period_type: str,
    month: int | None = None,
    quarter: int | None = None,
) -> BytesIO:
    q_months = {1: [1, 2, 3], 2: [4, 5, 6], 3: [7, 8, 9], 4: [10, 11, 12]}
    if period_type == "monthly" and month:
        months_range = [month]
        label = f"Tháng {month}/{year}"
    elif period_type == "quarterly" and quarter:
        months_range = q_months.get(quarter, [])
        label = f"Quý {quarter}/{year}"
    else:
        months_range = list(range(1, 13))
        label = f"Năm {year}"

    rev_p = [r for r in revenue if r["month"] in months_range]
    stu_p = [r for r in analytics.get("new_students_by_month", []) if r["month"] in months_range]
    total_rev = sum(r["total"] for r in rev_p)
    total_new = sum(r["count"] for r in stu_p)

    body = f"""<div class="hdr">
  <div>
    <div class="school">MÔ TÔ GIA THỊNH — THỐNG KÊ {html.escape(label.upper())}</div>
    <div class="sub">Xuất ngày {date.today().strftime("%d/%m/%Y")}</div>
  </div>
</div>
<div class="kpis">
  <div class="kpi"><div class="kpi-l">Tổng doanh thu</div><div class="kpi-v">{_v(total_rev)}</div></div>
  <div class="kpi"><div class="kpi-l">Học viên mới</div><div class="kpi-v">{total_new}</div></div>
  <div class="kpi"><div class="kpi-l">Tổng học viên</div><div class="kpi-v">{analytics.get("total_students", 0)}</div></div>
  <div class="kpi"><div class="kpi-l">Nợ quá hạn</div><div class="kpi-v red">{_v(analytics.get("overdue_amount", 0))}</div></div>
</div>"""

    if rev_p:
        rev_sorted = sorted(rev_p, key=lambda x: x["month"])
        rev_items = [(f"T{r['month']}", r["total"]) for r in rev_sorted]
        rev_chart = _svg_bar(rev_items, "#1677ff", lambda v: f"{v/1_000_000:.1f}tr")
        rows = "".join(f"<tr><td>{_MON[r['month']-1]}</td><td style='text-align:right'>{_v(r['total'])}</td></tr>" for r in rev_sorted)
        body += f"<h2>Doanh thu</h2>{rev_chart}<table><thead><tr><th>Tháng</th><th style='text-align:right'>Doanh thu</th></tr></thead><tbody>{rows}</tbody><tfoot><tr><td>Tổng cộng</td><td style='text-align:right'>{_v(total_rev)}</td></tr></tfoot></table>"

    if stu_p:
        stu_sorted = sorted(stu_p, key=lambda x: x["month"])
        stu_items = [(f"T{r['month']}", float(r["count"])) for r in stu_sorted]
        stu_chart = _svg_bar(stu_items, "#52c41a", lambda v: str(int(v)))
        rows = "".join(f"<tr><td>{_MON[r['month']-1]}</td><td style='text-align:right'>{r['count']}</td></tr>" for r in stu_sorted)
        body += f"<h2>Học viên đăng ký mới</h2>{stu_chart}<table><thead><tr><th>Tháng</th><th style='text-align:right'>Số lượng</th></tr></thead><tbody>{rows}</tbody><tfoot><tr><td>Tổng</td><td style='text-align:right'>{total_new}</td></tr></tfoot></table>"

    lic = analytics.get("students_by_license", [])
    if lic:
        lic_sorted = sorted(lic, key=lambda x: x["license_type"])
        lic_items = [(r["license_type"], float(r["count"])) for r in lic_sorted]
        lic_chart = _svg_bar(lic_items, "#722ed1", lambda v: str(int(v)))
        rows = "".join(f"<tr><td>Bằng {r['license_type']}</td><td style='text-align:right'>{r['count']}</td></tr>" for r in lic_sorted)
        body += f"<h2>Phân bổ loại bằng (tất cả thời gian)</h2>{lic_chart}<table><thead><tr><th>Loại bằng</th><th style='text-align:right'>Học viên</th></tr></thead><tbody>{rows}</tbody></table>"

    methods = analytics.get("payments_by_method", [])
    if methods:
        rows = "".join(f"<tr><td>{html.escape(_METHOD.get(r['phuong_thuc'], r['phuong_thuc']))}</td><td style='text-align:right'>{r['count']}</td><td style='text-align:right'>{_v(r['total'])}</td></tr>" for r in methods)
        body += f"<h2>Phương thức thanh toán — {year}</h2><table><thead><tr><th>Phương thức</th><th style='text-align:right'>Giao dịch</th><th style='text-align:right'>Tổng</th></tr></thead><tbody>{rows}</tbody></table>"

    body += f'<div class="foot"><span>Hệ thống quản lý Mô Tô Gia Thịnh</span><span>Ngày xuất: {date.today().strftime("%d/%m/%Y")}</span></div>'
    return _to_pdf(_doc(f"Thống kê {label}", body))


# ── Excel export ──────────────────────────────────────────────────────────────

def generate_analytics_excel(analytics: dict, revenue: list[dict], year: int) -> BytesIO:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1677FF")
    hdr_align = Alignment(horizontal="center")

    def _hdr(ws, cols: list[str]):
        ws.append(cols)
        for cell in ws[ws.max_row]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align

    wb = openpyxl.Workbook()

    # Revenue sheet
    ws = wb.active
    ws.title = "Doanh thu"
    _hdr(ws, ["Tháng", "Doanh thu (VNĐ)"])
    for r in sorted(revenue, key=lambda x: x["month"]):
        ws.append([_MON[r["month"] - 1], float(r["total"])])
    ws.append(["Tổng cộng", sum(float(r["total"]) for r in revenue)])
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20

    # New students
    ws2 = wb.create_sheet("Học viên mới")
    _hdr(ws2, ["Tháng", "Số học viên"])
    for r in sorted(analytics.get("new_students_by_month", []), key=lambda x: x["month"]):
        ws2.append([_MON[r["month"] - 1], r["count"]])
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 14

    # License types
    ws3 = wb.create_sheet("Loại bằng")
    _hdr(ws3, ["Loại bằng", "Số học viên"])
    for r in sorted(analytics.get("students_by_license", []), key=lambda x: x["license_type"]):
        ws3.append([f"Bằng {r['license_type']}", r["count"]])

    # Payment methods
    ws4 = wb.create_sheet("Thanh toán")
    _hdr(ws4, ["Phương thức", "Số giao dịch", "Tổng (VNĐ)"])
    for r in analytics.get("payments_by_method", []):
        ws4.append([_METHOD.get(r["phuong_thuc"], r["phuong_thuc"]), r["count"], float(r["total"])])
    ws4.column_dimensions["A"].width = 16
    ws4.column_dimensions["C"].width = 18

    # Lead sources
    ws5 = wb.create_sheet("Nguồn lead")
    _hdr(ws5, ["Nguồn", "Số lead"])
    src_labels = {"facebook": "Facebook", "walk_in": "Trực tiếp", "referral": "Giới thiệu", "zalo": "Zalo", "chatbot": "Chatbot", "other": "Khác"}
    for r in analytics.get("leads_by_source", []):
        ws5.append([src_labels.get(r["lead_source"], r["lead_source"]), r["count"]])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
